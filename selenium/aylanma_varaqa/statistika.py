from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment

excel_nomi = 'selenium/aylanma_varaqa/talaba_aylanma.xlsx'
# Excel faylini o'qing
wb = load_workbook(excel_nomi)

# Yangi list oching, agar mavjud bo'lmasa
list_name = "Statistika-tasdiqlanmagan"
if 'Statistika' not in wb.sheetnames:
    wb.create_sheet(title=list_name)
sheet = wb[list_name]
# C1 dan H1 gacha bo‘lgan kataklarni birlashtiramiz
sheet.merge_cells('C1:H1')
sheet.merge_cells('A1:A2')
sheet.merge_cells('B1:B2')

bold_font = Font(bold=True)

# Masalan, A1 dan A100 gacha bold qilish
for row in range(1, 20):
    for col in ['A', 'B']:
        sheet[f"{col}{row}"].font = bold_font


# Jirni  qilish
bold_font = Font(bold=True)
sheet['C1'].font = bold_font
sheet['A1'].font = bold_font
# Birlashtirilgan katakka matn yozamiz
sheet['C1'] = 'Tasdiqlanmagan'
sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

from openpyxl.styles import Font

bold_font = Font(bold=True)

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    sheet[f"{col}1"].font = bold_font
    sheet[f"{col}2"].font = bold_font

# Sarlavhalarni yozing
sheet['A1'] = 'Fakultet'
sheet['B1'] = 'Jami talaba'
sheet['C2'] = 'Registrator ofisi'
sheet['D2'] = 'Dekan'
sheet['E2'] = 'Marketing'
sheet['F2'] = 'Buxgalteriya'
sheet['G2'] = 'Kutubxona'
sheet['H2'] = 'Yotoqxona'

# Har bir fakultetdagi talabalar sonini va tasdiqlanmaganlarni hisoblash
fakultetlar = {}
tasdiqlanmagan = {
    'Registrator ofisi': {}, 'Dekan': {}, 'Marketing': {}, 'Buxgalteriya': {},
    'Kutubxona': {}, 'Yotoqxona': {}
}
for row in wb['Talabalar'].iter_rows(min_row=2, values_only=True):
    fakultet = row[6]  # G ustunida fakultet nomi
    rol = row[4]  # E ustunida rol
    ikonka = row[5]  # F ustunida ikonka (badge)
    if fakultet:
        # Fakultet bo'yicha umumiy talaba sonini hisoblash
        fakultetlar[fakultet] = fakultetlar.get(fakultet, 0) + 1
        if ikonka != "badge bg-green" and rol:
            for key in tasdiqlanmagan.keys():
                if key in rol:
                    tasdiqlanmagan[key][fakultet] = tasdiqlanmagan[key].get(fakultet, 0) + 1

# Ma'lumotlarni Excelga yozing
row_num = 3
for fakultet, soni in fakultetlar.items():
    sheet[f'A{row_num}'] = fakultet
    sheet[f'B{row_num}'] = (fakultetlar[fakultet])/ 6  # G ustunidagi son
    sheet[f'C{row_num}'] = tasdiqlanmagan['Registrator ofisi'].get(fakultet, 0)
    sheet[f'D{row_num}'] = tasdiqlanmagan['Dekan'].get(fakultet, 0)
    sheet[f'E{row_num}'] = tasdiqlanmagan['Marketing'].get(fakultet, 0)
    sheet[f'F{row_num}'] = tasdiqlanmagan['Buxgalteriya'].get(fakultet, 0)
    sheet[f'G{row_num}'] = tasdiqlanmagan['Kutubxona'].get(fakultet, 0)
    sheet[f'H{row_num}'] = tasdiqlanmagan['Yotoqxona'].get(fakultet, 0)
    row_num += 1

# Jami qatorini qo'shing
sheet[f'A{row_num}'] = 'Jami'
# Har bir ustun uchun umumiy sonni hisoblash
sheet[f'B{row_num}'] = (sum(fakultetlar.values()))/6  # Umumiy talaba soni
sheet[f'C{row_num}'] = sum(tasdiqlanmagan['Registrator ofisi'].values())
sheet[f'D{row_num}'] = sum(tasdiqlanmagan['Dekan'].values())
sheet[f'E{row_num}'] = sum(tasdiqlanmagan['Marketing'].values())
sheet[f'F{row_num}'] = sum(tasdiqlanmagan['Buxgalteriya'].values())
sheet[f'G{row_num}'] = sum(tasdiqlanmagan['Kutubxona'].values())
sheet[f'H{row_num}'] = sum(tasdiqlanmagan['Yotoqxona'].values())

# Faylni saqlang
wb.save(excel_nomi)