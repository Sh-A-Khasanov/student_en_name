from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from bs4 import BeautifulSoup
from selenium.webdriver.support import expected_conditions as EC
import time
import re
import os
from openpyxl import load_workbook, Workbook
import json
import math
import subprocess


                    # 2025 bu 2024-2025 o'quv yili hisoblanadi
uquv_yili = 2024    # 2024 bu 2024-2025 o'quv yili hisoblanadi
                    # 2023 bu 2023-2024 o'quv yili hisoblanadi



excel_nomi = 'selenium/aylanma_varaqa/talaba_aylanma.xlsx'

# Fayl borligini tekshirish
if os.path.exists(excel_nomi):
    wb = load_workbook(excel_nomi)
    sheet = wb.active
    excel_qator = sheet.max_row + 1
else:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Talabalar"

    # Sarlavhalar
    sheet['A1'] = 'T/R'
    sheet['B1'] = 'F.I.Sh.'
    sheet['C1'] = 'HEMIS ID'
    sheet['D1'] = 'Guruh'
    sheet['E1'] = 'Rol'
    sheet['F1'] = 'Ikonka'
    sheet['G1'] = 'Fakultet'
    sheet['H1'] = 'Izoh'
    sheet['I1'] = 'Tasdiq holati'


    excel_qator = 2  # Yangi fayl uchun 2-qator
    wb.save(excel_nomi)





wb = load_workbook(excel_nomi)
sheet = wb.active


driver = webdriver.Chrome()


# config.json faylini o‘qiymiz
with open("config.json", "r") as file:
    config = json.load(file)
otm_url = config["otm_url"]
driver.get(f"{otm_url}archive/circulation-sheet?ECirculationSheet%5B_education_year%5D={uquv_yili}")

# Login va parolni kiritish
login_value = config["login"]
parol_value = config["parol"]

login = driver.find_element(By.NAME, "FormAdminLogin[login]")
login.send_keys(login_value)
parol = driver.find_element(By.NAME, "FormAdminLogin[password]")
parol.send_keys(parol_value)
time.sleep(15)  # Sayt yuklanishini kutish va kirish tugmasini bosish uchun vaqt



jami_talabani_olish = driver.find_element(By.XPATH, "/html/body/div/div[2]/section[2]/div/div/div[2]/div[2]/span").text

# '1-50 / jami 21 286 ta'
match = re.search(r'jami\s([\d\s]+)\sta', jami_talabani_olish)
if match:
    jami_talaba = int(match.group(1).replace(" ", ""))
    print(f"jami_talaba : {jami_talaba}")  # 21286

total_page = math.ceil(jami_talaba/100)
print(f"total_page : {total_page}")
time.sleep(1)


# qaysi uquv_yili_uchun'ni ochish
wait = WebDriverWait(driver, 10)
fakultet = wait.until(EC.element_to_be_clickable((By.ID, "select2-ecirculationsheet-_department-container")))
fakultet.click()



# Ochilgan <li> elementlarini olish
li_elements = wait.until(EC.presence_of_all_elements_located(
    (By.CSS_SELECTOR, "ul#select2-ecirculationsheet-_department-results > li")))

fakultetlar = []  # (id, name) dan iborat ro'yxat

for li in li_elements:
    element_id = li.get_attribute("id")
    fakultet_name = li.text.strip()
    fakultet_id = element_id.split('-')[-1]
    fakultetlar.append((fakultet_id, fakultet_name))


for fakultet_id, fakultet_name in fakultetlar:
    try:
        # Har bir fakultet uchun umumiy sahifani aniqlash
        driver.get(f"{otm_url}archive/circulation-sheet?ECirculationSheet%5B_department%5D={fakultet_id}&ECirculationSheet%5B_education_year%5D={uquv_yili}")
        time.sleep(2)

        jami_talabani_olish = driver.find_element(By.XPATH, "/html/body/div/div[2]/section[2]/div/div/div[2]/div[2]/span").text
        match = re.search(r'jami\s([\d\s]+)\sta', jami_talabani_olish)
        if not match:
            continue  # hech nima topilmasa bu fakultetni o'tkazib yubor
        jami_talaba = int(match.group(1).replace(" ", ""))
        total_page = math.ceil(jami_talaba / 100)

        for page in range(1, total_page + 1):
            url = f"{otm_url}archive/circulation-sheet?page={page}&per-page=100&ECirculationSheet%5B_department%5D={fakultet_id}&ECirculationSheet%5B_education_year%5D={uquv_yili}"   
            driver.get(url)
            time.sleep(1)

            tr_elements = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            
            for tr_element in tr_elements:
                html = tr_element.get_attribute("outerHTML")
                soup = BeautifulSoup(html, 'html.parser')
                tr = soup.find('tr')

                name = tr.find_all('td')[1].contents[0].strip()
                group = tr.find_all('td')[3].find('p').text.strip()

                td_hemis = tr.find_all('td')[4]
                hemis_id = td_hemis.find(string=True, recursive=False).strip().split('/')[0]

                td_roles = tr.find_all('td')[5]
                roles = td_roles.select('span.badge, a > span.badge')

                # Tasdiq tugmasi <a> tegi
                tasdiq_qismi = tr.find_all('td')[6]  # 7-ustun

                a_tag = tasdiq_qismi.find('a', href=True)
                tasdiq_href = a_tag['href']

                # value=1 -> Tasdiqlangan, value=0 -> Tasdiqlanmagan
                if 'value=1' in tasdiq_href:
                    tasdiq_holati = "Tasdiqlangan"
                elif 'value=0' in tasdiq_href:
                    tasdiq_holati = "Tasdiqlanmagan"
                else:
                    tasdiq_holati = "Nomaʼlum"



                for badge in roles:
                    role_text = badge.get_text(strip=True)
                    class_list = badge.get('class', [])
                    class_str = ' '.join(class_list)
                    icon = badge.find('span', class_='fa')
                    icon_str = ' '.join(icon['class']) if icon else class_str
                    comment = badge.get('title', '').strip()

                    sheet[f'A{excel_qator}'] = excel_qator
                    sheet[f'B{excel_qator}'] = name
                    sheet[f'C{excel_qator}'] = hemis_id
                    sheet[f'D{excel_qator}'] = group
                    sheet[f'E{excel_qator}'] = role_text
                    sheet[f'F{excel_qator}'] = icon_str
                    sheet[f'G{excel_qator}'] = fakultet_name
                    sheet[f'H{excel_qator}'] = comment
                    sheet[f'I{excel_qator}'] = tasdiq_holati

                    excel_qator += 1

            print(f"{fakultet_name} => {page} / {total_page} sahifa saqlandi")
    except:
        print("Xatolik")
wb.save(excel_nomi)
time.sleep(2)
print("saqlandi")
driver.quit()
# asosiy kod tugaganidan keyin quyidagini qo‘shing
subprocess.run(["python", "selenium/aylanma_varaqa/statistika.py"])























