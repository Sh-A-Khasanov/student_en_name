from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import time
import re
import os
from openpyxl import load_workbook, Workbook
import json
import math


excel_nomi = 'talaba_aylanma.xlsx'

# Fayl borligini tekshirish
if os.path.exists(excel_nomi):
    wb = load_workbook(excel_nomi)
    sheet = wb.active
    excel_qator = sheet.max_row + 1
else:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Talabalar"

    # Sarlavhalar
    sheet['A1'] = 'T/R'
    sheet['B1'] = 'F.I.Sh.'
    sheet['C1'] = 'HEMIS ID'
    sheet['D1'] = 'Guruh'
    sheet['E1'] = 'Rol'
    sheet['F1'] = 'Ikonka'
    sheet['H1'] = 'Izoh'

    excel_qator = 2  # Yangi fayl uchun 2-qator
    wb.save(excel_nomi)





wb = load_workbook(excel_nomi)
sheet = wb.active


driver = webdriver.Chrome()


# config.json faylini o‘qiymiz
with open("config.json", "r") as file:
    config = json.load(file)
otm_url = config["otm_url"]
driver.get(f"{otm_url}archive/circulation-sheet?ECirculationSheet%5B_education_year%5D=2024")

# Login va parolni kiritish
login_value = config["login"]
parol_value = config["parol"]

login = driver.find_element(By.NAME, "FormAdminLogin[login]")
login.send_keys(login_value)
parol = driver.find_element(By.NAME, "FormAdminLogin[password]")
parol.send_keys(parol_value)
time.sleep(15)  # Sayt yuklanishini kutish va kirish tugmasini bosish uchun vaqt



jami_talabani_olish = driver.find_element(By.XPATH, "/html/body/div/div[2]/section[2]/div/div/div[2]/div[2]/span").text

# '1-50 / jami 21 286 ta'
match = re.search(r'jami\s([\d\s]+)\sta', jami_talabani_olish)
if match:
    jami_talaba = int(match.group(1).replace(" ", ""))
    print(f"jami_talaba : {jami_talaba}")  # 21286

total_page = math.ceil(jami_talaba/100)
print(f"total_page : {total_page}")
time.sleep(1)


for page in range(1, total_page+1):
# for page in range(1, 2):
    url = f"{otm_url}archive/circulation-sheet?page={page}&per-page=100&ECirculationSheet%5B_education_year%5D=2024"
    driver.get(url)
    time.sleep(2)

    tr_elements = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    

    for tr_element in tr_elements:
        html = tr_element.get_attribute("outerHTML")
        soup = BeautifulSoup(html, 'html.parser')
        tr = soup.find('tr')

        # Ism
        name = tr.find_all('td')[1].contents[0].strip()

        # Guruh
        group = tr.find_all('td')[3].find('p').text.strip()

        td_hemis = tr.find_all('td')[4]
        hemis_id = td_hemis.find(string=True, recursive=False).strip()
        hemis_id = hemis_id.split('/')[0]



        # Rollar (faqat tashqi badge'lar)
        td_roles = tr.find_all('td')[5]
        roles = td_roles.select('span.badge, a > span.badge')

        for badge in roles:
            role_text = badge.get_text(strip=True)
            class_list = badge.get('class', [])
            class_str = ' '.join(class_list)

            icon = badge.find('span', class_='fa')
            icon_str = ' '.join(icon['class']) if icon else class_str
            comment = badge.get('title', '').strip()
            # full_comment = f"{comment}" if comment else role_text
            full_comment = f"{comment}"

            # Har bir rol uchun alohida qatorda yozish
            sheet[f'A{excel_qator}'] = excel_qator
            sheet[f'B{excel_qator}'] = name
            sheet[f'C{excel_qator}'] = hemis_id
            sheet[f'D{excel_qator}'] = group
            sheet[f'E{excel_qator}'] = role_text
            sheet[f'F{excel_qator}'] = icon_str
            sheet[f'H{excel_qator}'] = full_comment
            excel_qator += 1

    print(f"{page} / {total_page} ta sahifa saqlandi")
    wb.save(excel_nomi)
driver.quit()
